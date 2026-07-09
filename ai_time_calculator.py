from flask import Flask, request, send_file, render_template_string
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from werkzeug.utils import secure_filename
from spellchecker import SpellChecker
import os
import re

try:
    import pythoncom
    import win32com.client as win32
    WIN32_AVAILABLE = True
except Exception:
    WIN32_AVAILABLE = False

app = Flask(__name__)
spell = SpellChecker()

RED = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

HTML = """
<!DOCTYPE html>
<html>
<head>
<title>AI Excel Report Checker</title>
<style>
body{font-family:Arial;background:#f4f6f8;text-align:center;padding-top:70px}
.box{background:white;width:600px;margin:auto;padding:35px;border-radius:12px;box-shadow:0 0 12px #ccc}
button{background:#1f4e79;color:white;border:none;padding:14px 28px;border-radius:6px;cursor:pointer;font-weight:bold}
button:hover{background:#153552}
input{margin:15px;width:85%;padding:8px}
</style>
</head>
<body>
<div class="box">
<h2>AI Excel Report Checker</h2>
<form action="/upload" method="post" enctype="multipart/form-data">
<input type="file" name="excel_file" accept=".xlsx" required>
<br>
<button type="submit">Upload & Check</button>
</form>
</div>
</body>
</html>
"""

PRESENT_WORDS = [
    "working", "preparing", "creating", "updating", "verifying",
    "discussing", "searching", "contacting", "ongoing",
    "developing", "designing", "checking", "making", "calling"
]

PAST_WORDS = [
    "completed", "finished", "prepared", "created", "updated",
    "verified", "discussed", "searched", "contacted", "shared",
    "sent", "checked", "made", "developed", "designed", "called"
]

COMPLETED_STATUS = ["completed", "done", "finished"]
WIP_STATUS = ["work in progress", "work in process", "wip", "pending"]


def is_header_row(ws, row):
    values = [str(ws.cell(row, col).value or "").strip().lower() for col in range(1, 7)]
    values = [v.replace("s no", "s.no").replace("sno", "s.no") for v in values]

    return (
        values[0] == "s.no"
        and values[1] == "task"
        and values[2] == "status"
        and values[3] == "timing"
        and values[4] == "date of completion"
        and values[5] == "remarks"
    )


def is_info_row(ws, row):
    text = " ".join(str(ws.cell(row, col).value or "").lower() for col in range(1, 8))
    info_words = ["date", "name", "department", "designation"]
    return sum(1 for word in info_words if word in text) >= 2


def detect_tense(text):
    text = str(text or "").lower()

    if any(re.search(rf"\b{re.escape(word)}\b", text) for word in PRESENT_WORDS):
        return "Present tense"

    if any(re.search(rf"\b{re.escape(word)}\b", text) for word in PAST_WORDS):
        return "Past tense"

    return "Tense unclear"


def get_wrong_words(status, remark):
    status_lower = str(status or "").strip().lower()
    remark_lower = str(remark or "").lower()
    wrong_words = []

    if status_lower in COMPLETED_STATUS:
        for word in PRESENT_WORDS:
            if re.search(rf"\b{re.escape(word)}\b", remark_lower):
                wrong_words.append(word)

    elif status_lower in WIP_STATUS:
        for word in PAST_WORDS:
            if re.search(rf"\b{re.escape(word)}\b", remark_lower):
                wrong_words.append(word)

    return wrong_words


def correct_remark(status, remark):
    text = str(remark or "").strip()

    if text == "":
        return ""

    status_lower = str(status or "").strip().lower()

    completed_replace = {
        "developing": "developed",
        "creating": "created",
        "preparing": "prepared",
        "checking": "checked",
        "updating": "updated",
        "verifying": "verified",
        "searching": "searched",
        "contacting": "contacted",
        "discussing": "discussed",
        "designing": "designed",
        "calling": "called",
        "making": "made",
        "working": "worked"
    }

    wip_replace = {
        "developed": "developing",
        "created": "creating",
        "prepared": "preparing",
        "checked": "checking",
        "updated": "updating",
        "verified": "verifying",
        "searched": "searching",
        "contacted": "contacting",
        "discussed": "discussing",
        "designed": "designing",
        "called": "calling",
        "made": "making",
        "completed": "working on",
        "finished": "working on",
        "done": "working on"
    }

    if status_lower in COMPLETED_STATUS:
        for old, new in completed_replace.items():
            text = re.sub(rf"\b{re.escape(old)}\b", new, text, flags=re.IGNORECASE)

    elif status_lower in WIP_STATUS:
        for old, new in wip_replace.items():
            text = re.sub(rf"\b{re.escape(old)}\b", new, text, flags=re.IGNORECASE)

        if "progress" not in text.lower() and "ongoing" not in text.lower():
            text = text.rstrip(".") + " and the work is in progress"

    text = re.sub(r"\bi\b", "I", text)
    text = re.sub(r"\bim\b", "I am", text, flags=re.IGNORECASE)
    text = re.sub(r"\bdont\b", "do not", text, flags=re.IGNORECASE)
    text = re.sub(r"\bcant\b", "cannot", text, flags=re.IGNORECASE)

    text = text[0].upper() + text[1:]

    if text[-1] not in ".!?":
        text += "."

    return text


def check_spelling(text):
    issues = []
    suggestions = []

    ignore_words = {
        "sk", "hr", "it", "wip", "whatsapp", "google",
        "sasi", "sir", "sapientia", "excel", "gmail",
        "sri", "kalaivani", "ppt", "pdf", "ceo",

        # person names
        "sanjay", "kumar", "vignesh", "surabi", "sashik",
        "priya", "divya", "karthik", "arun", "surya",
        "naveen", "ravi", "rajesh", "ramesh", "suba",
        "subash", "meena", "ramya", "deepa", "anitha",
        "anita", "lakshmi", "vinoth", "vijay", "ajay",
        "hari", "balaji", "dinesh", "ganesh", "mohan",
        "kavin", "kavya", "keerthi", "nandhini",
        "suresh", "mahesh", "raja", "kaviya", "swetha",
        "shweta", "shalini", "saranya", "janani", "pooja",
        "monisha", "harini", "abi", "abirami", "gokul",
        "prakash", "bharath", "vishnu", "manoj", "sathish",
        "sakthi", "muthu", "yuvaraj", "venkat", "senthil"
    }

    manual_corrections = {
        "developng": "developing",
        "develping": "developing",
        "devloping": "developing",
        "creting": "creating",
        "craeting": "creating",
        "updting": "updating",
        "updatng": "updating",
        "verifing": "verifying",
        "verifyng": "verifying",
        "discusssion": "discussion",
        "discusson": "discussion",
        "discusion": "discussion",
        "prepairing": "preparing",
        "preparingg": "preparing",
        "serching": "searching",
        "searcing": "searching",
        "contacing": "contacting",
        "contactng": "contacting",
        "completd": "completed",
        "complet": "complete",
        "completingg": "completing",
        "recived": "received",
        "recieve": "receive",
        "recieved": "received",
        "messege": "message",
        "mesage": "message",
        "whastapp": "whatsapp",
        "watsapp": "whatsapp",
        "candidte": "candidate",
        "candiate": "candidate",
        "reprot": "report",
        "reprt": "report",
        "repot": "report",
        "clinet": "client",
        "offcie": "office",
        "ofice": "office",
        "postr": "poster",
        "desgin": "design",
        "desinging": "designing",
        "varified": "verified",
        "cheking": "checking",
        "serched": "searched",
        "shred": "shared",
        "sharring": "sharing",
        "selcted": "selected",
        "agnt": "agent",
        "carrer": "career",
        "guidence": "guidance"
    }

    words = re.findall(r"[A-Za-z]+", str(text or ""))

    clean_words = []

    for word in words:
        lower_word = word.lower()

        if len(lower_word) <= 2:
            continue

        if lower_word in ignore_words:
            continue

        if lower_word in manual_corrections:
            issues.append(f"Spelling mistake: '{word}'")
            suggestions.append(f"Correct spelling: {word} → {manual_corrections[lower_word]}")
            continue

        clean_words.append(lower_word)

    misspelled = spell.unknown(clean_words)

    for word in misspelled:
        correct_word = spell.correction(word)

        if correct_word and correct_word != word:
            issues.append(f"Spelling mistake: '{word}'")
            suggestions.append(f"Correct spelling: {word} → {correct_word}")

    return issues, suggestions


def check_task_case(task):
    issues = []

    if not task:
        issues.append("Task missing")
        return issues

    for word in str(task).split():
        clean = re.sub(r"[^A-Za-z]", "", word)

        if clean and clean[0].islower():
            issues.append(f"Task title casing error: '{word}' should start with capital letter")

    return issues


def word_set(text):
    return set(re.findall(r"[a-zA-Z]+", str(text).lower()))


def check_remark(task, status, remark):
    issues = []
    suggestions = []

    task = str(task or "").strip()
    remark = str(remark or "").strip()

    if "lunch" in task.lower() and remark in ["", "-", "---", "------"]:
        return issues, ["Correct"]

    if remark == "":
        issues.append("Remark missing")
        suggestions.append("Please enter a clear remark.")
        return issues, suggestions

    spelling_issues, spelling_suggestions = check_spelling(remark)

    if spelling_issues:
        issues.extend(spelling_issues)
        suggestions.extend(spelling_suggestions)

    wrong_words = get_wrong_words(status, remark)

    if wrong_words:
        issues.append("Wrong tense word found in remark")

    corrected = correct_remark(status, remark)

    if corrected != remark:
        suggestions.append("Correct remark: " + corrected)

    suggestions.append("Detected tense: " + detect_tense(remark))

    if len(remark.split()) < 3:
        issues.append("Remark is too short or vague")

    if remark and remark[0].islower():
        issues.append("Remark should start with capital letter")

    if remark.isupper() and len(remark) > 5:
        issues.append("Remark should not be fully uppercase")

    vague_words = ["done", "ok", "okay", "completed", "work", "same", "nil", "nothing"]

    if remark.lower() in vague_words:
        issues.append("Remark is unclear")

    ignore_words = {
        "the", "a", "an", "and", "or", "to", "for", "with",
        "in", "on", "of", "is", "was", "were", "work", "task",
        "done", "completed", "about", "regarding", "i", "we"
    }

    task_words = word_set(task) - ignore_words
    remark_words = word_set(remark) - ignore_words

    if task_words and remark not in ["-", "---", "------", "NA", "N/A"]:
        if not task_words.intersection(remark_words):
            issues.append("Task and remark are not matching")

    if not suggestions:
        suggestions.append("Correct")

    return issues, suggestions


def check_excel_openpyxl(input_path, output_path):
    wb = load_workbook(input_path)

    for ws in wb.worksheets:
        header_row = None

        for row in range(1, ws.max_row + 1):
            if is_header_row(ws, row):
                header_row = row
                break

        if header_row is None:
            continue

        check_status_col = 7
        reason_col = 8
        suggestion_col = 9

        ws.cell(header_row, check_status_col).value = "Check Status"
        ws.cell(header_row, reason_col).value = "Reason Summary"
        ws.cell(header_row, suggestion_col).value = "Suggestions"

        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 65
        ws.column_dimensions["I"].width = 85

        for col in range(1, suggestion_col + 1):
            ws.cell(header_row, col).font = Font(bold=True)
            ws.cell(header_row, col).alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="center"
            )

        for row in range(header_row + 1, ws.max_row + 1):
            if is_header_row(ws, row) or is_info_row(ws, row):
                continue

            task_cell = ws.cell(row, 2)
            status_cell = ws.cell(row, 3)
            completion_cell = ws.cell(row, 5)
            remark_cell = ws.cell(row, 6)

            task = task_cell.value
            status = status_cell.value
            completion = completion_cell.value
            remark = remark_cell.value

            if not task and not status and not completion and not remark:
                continue

            issues = []

            case_issues = check_task_case(task)

            if case_issues:
                issues.extend(case_issues)
                task_cell.fill = YELLOW

            status_lower = str(status or "").strip().lower()

            if status_lower in WIP_STATUS:
                if completion is None or str(completion).strip() in ["", "-", "---", "------"]:
                    issues.append("Completion date missing for WIP/Pending task")
                    completion_cell.fill = YELLOW

            remark_issues, suggestions = check_remark(task, status, remark)

            if remark_issues:
                issues.extend(remark_issues)
                remark_cell.fill = YELLOW

            ws.cell(row, suggestion_col).value = "\n".join(suggestions)

            if issues:
                ws.cell(row, check_status_col).value = "Error Found"
                ws.cell(row, reason_col).value = "\n".join(issues)
                ws.cell(row, check_status_col).fill = RED
                ws.cell(row, reason_col).fill = RED
                ws.row_dimensions[row].height = 60
            else:
                ws.cell(row, check_status_col).value = "Correct"
                ws.cell(row, check_status_col).fill = GREEN
                ws.cell(row, reason_col).value = "No issue found"
                ws.row_dimensions[row].height = 45

            ws.cell(row, reason_col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, suggestion_col).alignment = Alignment(wrap_text=True, vertical="top")
            remark_cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)


def color_word_red(cell, start_pos, length):
    try:
        cell.Characters(Start=start_pos, Length=length).Font.Color = 255
        return True
    except Exception:
        pass

    try:
        cell.Characters(start_pos, length).Font.Color = 255
        return True
    except Exception:
        pass

    try:
        chars = cell.GetCharacters(start_pos, length)
        chars.Font.Color = 255
        return True
    except Exception:
        pass

    return False


def highlight_wrong_words_excel(output_path):
    if not WIN32_AVAILABLE:
        return

    pythoncom.CoInitialize()
    excel = None
    wb = None

    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(os.path.abspath(output_path))

        for ws in wb.Worksheets:
            used_rows = ws.UsedRange.Rows.Count

            for row in range(1, used_rows + 1):
                status = ws.Cells(row, 3).Value
                remark = ws.Cells(row, 6).Value

                if not status or not remark:
                    continue

                wrong_words = get_wrong_words(status, remark)

                if not wrong_words:
                    continue

                text = str(remark)
                cell = ws.Cells(row, 6)

                for word in wrong_words:
                    for match in re.finditer(rf"\b{re.escape(word)}\b", text, flags=re.IGNORECASE):
                        start_pos = match.start() + 1
                        length = match.end() - match.start()
                        success = color_word_red(cell, start_pos, length)

                        if not success:
                            cell.Font.Color = 255

        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()

    except Exception:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass

        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass

    finally:
        pythoncom.CoUninitialize()


def check_excel(input_path, output_path):
    check_excel_openpyxl(input_path, output_path)
    highlight_wrong_words_excel(output_path)


@app.route("/")
def home():
    return render_template_string(HTML)


@app.route("/upload", methods=["POST"])
def upload():
    file = request.files.get("excel_file")

    if file is None or file.filename == "":
        return "No file selected", 400

    if not file.filename.lower().endswith(".xlsx"):
        return "Please upload only .xlsx Excel file", 400

    os.makedirs("uploads", exist_ok=True)
    os.makedirs("outputs", exist_ok=True)

    filename = secure_filename(file.filename)
    time_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

    input_path = os.path.abspath(os.path.join("uploads", f"{time_id}_{filename}"))
    output_path = os.path.abspath(os.path.join("outputs", f"AI_Checked_{time_id}_{filename}"))

    file.save(input_path)

    try:
        check_excel(input_path, output_path)
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        return f"Error while processing Excel: {e}", 500


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)